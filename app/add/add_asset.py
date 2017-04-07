from app import app, db, registry
from app.models import Site, AssetType, FunctionalDescriptor, LoggedEntity, Asset, PointType, AssetPoint, Algorithm
from flask import request, render_template, url_for, redirect, flash, send_file, make_response, jsonify
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from io import BytesIO
from xml.etree.ElementTree import ElementTree, Element, SubElement

# page to add an asset to the site
@app.route('/site/<sitename>/add')
def add_asset_input(sitename):
    site = Site.query.filter_by(name=sitename).one()
    asset_types = AssetType.query.all()
    return render_template('add_asset.html', site=site, asset_types=asset_types)

# return list of process functions through AJAX
@app.route('/site/<sitename>/add/_function')
def return_functions(sitename):
    asset_type = AssetType.query.filter_by(name=request.args['type']).one()
    function_names = [function.name for function in asset_type.functions]
    return jsonify(function_names)

# return list of loggedentities through AJAX
@app.route('/site/<sitename>/add/_loggedentity')
def return_loggedentities(sitename):
    site = Site.query.filter_by(name=sitename).one()

    # get database session for this site
    session = registry.get(site.db_key)

    if not session is None:
        logs = session.query(LoggedEntity).filter_by(type='trend.ETLog').all()
        log_paths = [log.path for log in logs]
        session.close()
    else:
        log_paths = []

    return jsonify(log_paths)

# return list of points through AJAX
@app.route('/site/<sitename>/add/_point')
def return_points(sitename):
    points = PointType.query.all()
    point_names = [point.name for point in points]
    return jsonify(point_names)

# return list of algorithms through AJAX
@app.route('/site/<sitename>/add/_algorithm')
def return_algorithms(sitename):
    algorithms = Algorithm.query.all()
    algorithm_names = [algorithm.descr for algorithm in algorithms]
    return jsonify(algorithm_names)

# process an asset addition
@app.route('/site/<sitename>/add/_submit', methods=['POST'])
def add_asset_submit(sitename):

    site = Site.query.filter_by(name=sitename).one()

    # get database session for this site
    session = registry.get(site.db_key)

    # create asset with 0 health
    type = AssetType.query.filter_by(name=request.form['type']).one()
    asset = Asset(type=type, name=request.form['name'], location=request.form['location'], group=request.form['group'], priority=request.form['priority'], site=site, health=0)
    db.session.add(asset)

    # TODO: need a better system of reading in values than string-matching point1 and log1
    # generate points
    for i in range(1, len(PointType.query.all()) + 1):
        point_type_name = request.form.get('point' + str(i))
        if not point_type_name is None:

            point_type = PointType.query.filter_by(name=point_type_name).one()
            point = AssetPoint(type=point_type, name=point_type_name)

            # assign the log id to the point
            log_path = request.form.get('log' + str(i))
            if not log_path is None and not session is None:
                log = session.query(LoggedEntity).filter_by(path=log_path).one()
                point.loggedentity_id = log.id

            asset.points.append(point)

    # set process functions
    function_list = request.form.getlist('function')
    for function_name in function_list:
        function = FunctionalDescriptor.query.filter_by(name=function_name).one()
        asset.functions.append(function)

    # set excluded algorithms
    inclusions = []
    included_list = request.form.getlist('algorithm')
    for algorithm_descr in included_list:
        inclusions.append(Algorithm.query.filter_by(descr=algorithm_descr).one())
    exclusions = set(Algorithm.query.all()) - set(inclusions)
    asset.exclusions.extend(exclusions)

    db.session.commit()
    session.close()

    return redirect(url_for('asset_list', sitename=sitename))

# generate a downloadable Excel template for uploading assets
@app.route('/site/<sitename>/add/_download')
def add_asset_download(sitename):

    wb = Workbook()
    ws_input = wb.worksheets[0]

    # generate input page
    ws_input.title = "Input"
    ws_input['A1'] = "Name"
    ws_input['B1'] = "Location"
    ws_input['C1'] = "Group"
    ws_input['D1'] = "Type"
    ws_input['E1'] = "Priority"

    # generate page with all the options (for data validation)
    ws_options = wb.create_sheet("Options")
    x = 1
    for asset_type in AssetType.query.all():
        ws_options.cell(column=x, row=1).value = asset_type.name
        x += 1
    cols = len(tuple(ws_options.columns))
    wb.create_named_range("Types", ws_options, '$A$1:${}$1'.format(get_column_letter(cols)))

    # apply data validation to input page
    type_dv = DataValidation(type='list', formula1='Types', allow_blank=True)
    priority_dv = DataValidation(type='whole', operator='between', formula1=0, formula2=9)
    ws_input.add_data_validation(type_dv)
    ws_input.add_data_validation(priority_dv)
    type_dv.ranges.append('D2:D1000')
    priority_dv.ranges.append('E2:E1000')

    # save file
    out = BytesIO(save_virtual_workbook(wb))

    # prevent browser from caching the download
    response = make_response(send_file(out, attachment_filename='Template.xlsx', as_attachment=True))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return response

# process an asset addition via upload
@app.route('/site/<sitename>/add/confirm', methods=['POST'])
def add_asset_upload(sitename):

    site = Site.query.filter_by(name=sitename).one()

    # check if file was uploaded
    if 'file' not in request.files:
        flash('No file selected')
        return redirect(url_for('add_asset_input', sitename=sitename))
    file = request.files['file']
    if file.filename =='':
        flash('No file selected')
        return redirect(url_for('add_asset_input', sitename=sitename))

    # check if file is correct type
    if not (file and ('.' in file.filename) and (file.filename.rsplit('.',1)[1].lower() in ['xlsx','xlsm'])):
        flash('File must be xlsx/xlsm')
        return redirect(url_for('add_asset_input', sitename=sitename))

    wb = load_workbook(file)
    ws = wb.worksheets[0]

    # generate asset for each non-blank row in the worksheet
    asset_list = []
    for row in tuple(ws.rows)[1:]:
        # skip if not all required information is present
        if not row[0].value is None and not row[3].value is None and not row[4].value is None:
            name = row[0].value
            location = row[1].value
            if location is None:
                location = ""
            group = row[2].value
            if group is None:
                group = ""
            type_name = row[3].value
            priority = row[4].value
            asset_type = AssetType.query.filter_by(name=type_name).one()
            asset = Asset(name=name, location=location, group=group, type=asset_type, priority=priority)
            asset_list.append(asset)

    return render_template('add_asset_confirm.html', assets=asset_list, site=site)

# JSON confirmation for an asset upload
# redirection after form submit occurs from within the HTML form
# this function is only triggered by a javascript AJAX POST request (easiest way to post a JSON object to group data together)
# which is triggered when the submit button is clicked
@app.route('/site/<sitename>/add/_upload_confirm', methods=['POST'])
def add_asset_confirm(sitename):
    site = Site.query.filter_by(name=sitename).one()
    assets_json = request.get_json(force=True)
    # returns a dict with each 'value' being the data for one asset, so iterate through the values
    for asset_json in assets_json.values():
        asset_type = AssetType.query.filter_by(name=asset_json['type']).one()
        asset = Asset(name=asset_json['name'], location=asset_json['location'], group=asset_json['group'], type=asset_type, priority=asset_json['priority'], health=0, site=site)
        db.session.add(asset)
        print(asset)
    db.session.commit()


# OLD CODE FOR GENERATING XML IMPORT - LEFT HERE TO BE REUSED LATER

# # process an asset addition via upload
# @app.route('/site/<sitename>/add/_upload', methods=['POST'])
# def add_asset_upload(sitename):

#     site = Site.query.filter_by(name=sitename).one()

#     # check if file was uploaded
#     if 'file' not in request.files:
#         flash('No file part')
#         return redirect(url_for('add_asset_input', sitename=sitename))
#     file = request.files['file']
#     if file.filename =='':
#         flash('No file selected')
#         return redirect(url_for('add_asset_input', sitename=sitename))

#     # check if file is correct type
#     if not (file and ('.' in file.filename) and (file.filename.rsplit('.',1)[1].lower() in ['xlsx','xlsm'])):
#         flash('File must be xlsx/xlsm')
#         return redirect(url_for('add_asset_input', sitename=sitename))

#     wb = load_workbook(file)
#     ws = wb.worksheets[0]

#     # generate asset for each non-blank row in the worksheet
#     asset_list = []
#     for row in tuple(ws.rows)[1:]:
#         if not row[0].value is None and not row[3].value is None and not row[4].value is None and not row[5].value is None:
#             name = row[0].value
#             location = row[1].value
#             if location is None:
#                 location = ""
#             group = row[2].value
#             if group is None:
#                 group = ""
#             type_name = row[3].value
#             subtype_name = row[4].value
#             priority = row[5].value
#             asset_type = AssetType.query.filter_by(name=type_name).one()
#             subtype = FunctionalDescriptor.query.filter_by(name=subtype_name, type=asset_type).one()
#             asset = Asset(name=name, location=location, group=group, subtype=subtype, priority=priority, site=site, health=0)

#             db.session.add(asset)
#             asset_list.append(asset)

#     # generate points
#     for subtype_point in subtype.points:
#         point = AssetPoint(type=subtype_point.type, asset=asset, name=subtype_point.type.name)

#         # set destination folder
#         point.loggedentity_path = "/Server 1/Medusa/Trends/{}/{} Extended".format(asset.name, point.name)

#         db.session.add(point)

#     db.session.commit()

#     # generate xml import to SBO
#     object_set = Element('ObjectSet')
#     object_set.attrib = {'ExportMode':'Standard', 'Version':'1.8.1.87', 'Note':'TypesFirst'}
#     exported_objects = SubElement(object_set, 'ExportedObjects')
#     for asset in asset_list:
#         oi_folder = SubElement(exported_objects, 'OI')
#         oi_folder.attrib = {'NAME':asset.name, 'TYPE':'system.base.Folder'}
#         for point in asset.points:
#             oi_trend = SubElement(oi_folder, 'OI')
#             oi_trend.attrib = {'NAME':'{} Extended'.format(point.name), 'TYPE':'trend.ETLog'}
#             pi = SubElement(oi_trend, 'PI')
#             pi.attrib = {'Name':'IncludeInReports', 'Value':'1'}

#     # save file
#     out = BytesIO()
#     tree = ElementTree(object_set)
#     tree.write(out, encoding='utf-8', xml_declaration=True)
#     out.seek(0)

#     # prevent browser from caching the download
#     response = make_response(send_file(out, attachment_filename='Import.xml', as_attachment=True))
#     response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
#     return response